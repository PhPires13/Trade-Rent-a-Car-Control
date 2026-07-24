"""Exportação Excel/CSV para envio à contabilidade (decisão nº 18)."""

import csv
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font


def planilha_xlsx(nome_arquivo, titulo, colunas, linhas):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    ws.append(colunas)
    for celula in ws[1]:
        celula.font = Font(bold=True)
    for linha in linhas:
        ws.append(list(linha))
    for indice, coluna in enumerate(colunas, 1):
        largura = max(
            [len(str(coluna))]
            + [len(str(linha[indice - 1])) for linha in linhas if len(linha) >= indice]
        )
        ws.column_dimensions[ws.cell(row=1, column=indice).column_letter].width = min(
            largura + 2, 60
        )
    buffer = BytesIO()
    wb.save(buffer)
    resposta = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resposta["Content-Disposition"] = f'attachment; filename="{nome_arquivo}.xlsx"'
    return resposta


def planilha_csv(nome_arquivo, colunas, linhas):
    resposta = HttpResponse(content_type="text/csv")
    resposta["Content-Disposition"] = f'attachment; filename="{nome_arquivo}.csv"'
    writer = csv.writer(resposta)
    writer.writerow(colunas)
    writer.writerows(linhas)
    return resposta


def exportar(formato, nome_arquivo, titulo, colunas, linhas):
    if formato == "xlsx":
        return planilha_xlsx(nome_arquivo, titulo, colunas, linhas)
    return planilha_csv(nome_arquivo, colunas, linhas)
